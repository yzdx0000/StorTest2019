#!/usr/bin/python
# -*- encoding=utf8 -*-

"""
Author: baorb
Description: 混合故障脚本
"""
import os
import re
import sys
import time
import json
import random
import signal
import logging
import openpyxl
import traceback
import subprocess
from optparse import OptionParser
from multiprocessing import Process, Manager

"""
Case_Lst: 要运行的用例列表, 成员是dic, {'row':1,'casename':'case_001','nums':1,'fault':[1,2],'nodenum':3,'node_parity_num':1,
                                   'disk_parity_num':2,'interval':'1-300','result':'succeed','result_col_idx':9}
"""
Case_Lst = []
Case_File = u'case_list.xlsx'        # 用例表文件名

wait_times = {'down_disk': [1, 600], 'del_disk': [300, 600],
              'down_net': [1, 600], 'del_node': [600, 1200],
              'down_node': [1, 300]}

File_Path = os.path.split(os.path.realpath(__file__))[0]

global NODE_IP_LST              # 集群节点的管理ip,列表
global NODE_DATA_IP_LST         # 集群节点的数据ip,列表
global CORESTOP                 # 遇到core是否退出
global NUMBERS                  # 故障循环次数

CHECKBADOBJ_WAIT_TIME = 180    # 故障完成后到检查坏对象的等待时间, 单位:s

Process_Dic = {'zk': 'zk',
               'oJmgs': '/home/parastor/bin/oJmgs',
               'oMgcd': '/home/parastor/bin/oMgcd',
               'oPara': '/home/parastor/bin/oPara',
               'oStor': '/home/parastor/bin/oStor',
               'oJob': '/home/parastor/bin/oJob',
               'oRole': '/home/parastor/bin/oRole',
               'oCnas': '/home/parastor/bin/oCnas',
               'oMgcd_client': '/cliparastor/bin/oMgcd_client'}

"""vdbench 参数设置"""
Vdb_Path = '/home/vdbench50406'
System_Ips = ('10.2.40.95', '10.2.40.96')
Anchor_Path = '/mnt/volume1/'

"""****************************** 故障分类 ******************************"""
Fault_Type = {'node': (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13),
              'net': (15, 16, 17, 18, 21, 22),
              'disk': (23, 24, 25, 26),
              'pro': (29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40)
              }

"""****************************** vdbench ******************************"""


class Vdbenchrun():
    """
    vdbench运行
    """
    def __init__(self, vdb_path, system_ips, anchor_path):
        self.vdb_path = vdb_path
        self.system_ips = system_ips
        self.anchor_path = anchor_path
        self.depth = 2
        self.width = 5
        self.files = 3000
        self.size = '(4k,35,128k,30,1m,20,10m,10,100m,5)'
        # self.size = '(128k,60,256k,40)'
        self.threads = 20
        self.xfersize = '(4k,30,32k,30,128k,30,1m,10)'
        self.elapsed = 300000
        self.config_file = os.path.join(File_Path, 'vdb_test')
        self._set_file()
        return

    def run_vdb(self):
        vdbench_exe_file = os.path.join(self.vdb_path, 'vdbench')
        vdbench_output_path = os.path.join(File_Path, 'output')
        cmd = "sh %s -f %s -o %s -v" % (vdbench_exe_file, self.config_file, vdbench_output_path)
        process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        """
        while True:
            line = process.stdout.readline()
            if '' == line:
                break
            print(line.rstrip())
        """
        stdout, stderr = process.communicate()
        retcode = process.poll()
        return retcode, stdout, stderr

    def _set_file(self):
        file_content_lst = []
        line_str = "data_errors=1"
        file_content_lst.append(line_str)

        line_str = ""
        file_content_lst.append(line_str)

        """hd部分"""
        line_str = "hd=default,vdbench=%s,shell=ssh,user=root" % self.vdb_path
        file_content_lst.append(line_str)

        i = 0
        if isinstance(self.system_ips, str):
            system_ips = (self.system_ips)
        elif isinstance(self.system_ips, (list, tuple)):
            system_ips = self.system_ips
        else:
            raise Exception("System_Ips is not right!!!")
        for system_ip in system_ips:
            i += 1
            line_str = "hd=h%d,system=%s" % (i, system_ip)
            file_content_lst.append(line_str)

        line_str = ""
        file_content_lst.append(line_str)

        """fsd部分"""
        line_str = "fsd=default,depth=%d,width=%d,files=%d,size=%s" % (self.depth, self.width, self.files, self.size)
        file_content_lst.append(line_str)

        for i in range(len(system_ips)):
            i += 1
            line_str = "fsd=fsd%d,anchor=%s/vdbench%d" % (i, self.anchor_path, i)
            file_content_lst.append(line_str)

        line_str = ""
        file_content_lst.append(line_str)

        """fwd部分"""
        line_str = "fwd=default,threads=%d,xfersize=%s" % (self.threads, self.xfersize)
        file_content_lst.append(line_str)

        for i in range(len(system_ips)):
            i += 1
            line_str = "fwd=fwd%d,fsd=fsd%d,hd=h%d" % (i, i, i)
            file_content_lst.append(line_str)

        line_str = ""
        file_content_lst.append(line_str)

        """rd部分"""
        line_str = "rd=default,fwdrate=1000,elapsed=%d,interval=1,warmup=20,pause=20" % self.elapsed
        file_content_lst.append(line_str)

        for i in range(20):
            i += 1
            line_str = "rd=rd%d,fwd=fwd*,format=yes,forrdpct=50,forseekpct=50" % i
            file_content_lst.append(line_str)

        """写入配置文件"""
        file_content = '\n'.join(file_content_lst)

        create_file_name = os.path.join(self.config_file)
        with open(create_file_name, 'w') as f:
            f.write(file_content)
        return


def run_vdbench(vdb_fail_flag):
    obj_vdb = Vdbenchrun(Vdb_Path, System_Ips, Anchor_Path)
    while True:
        rc, output, outerr = obj_vdb.run_vdb()
        if rc != 0:
            vdb_fail_flag.value = 1
            logging.error("vdbench failed!!!")
            logging.error(outerr)
            raise Exception("vdbench failed!!!")

"""****************************** 初始化日志 ******************************"""


def log_init():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 日志初始化
    :return: 
    """
    file_name = os.path.splitext(os.path.basename(__file__))[0]
    now_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
    file_name = now_time + '_' + file_name + '.log'
    file_name = os.path.join(File_Path, file_name)
    print file_name
    logging.basicConfig(level=logging.DEBUG,
                        format='[%(levelname)s][%(asctime)s]%(lineno)d:  %(message)s',
                        datefmt='%y-%m-%d %H:%M:%S',
                        filename=file_name,
                        filemode='a')

    console = logging.StreamHandler()
    console.setLevel(logging.DEBUG)
    formatter = logging.Formatter('[%(levelname)s][%(asctime)s]   %(message)s', '%y-%m-%d %H:%M:%S')
    console.setFormatter(formatter)
    logging.getLogger().addHandler(console)
    return


"""****************************** 公共函数 ******************************"""


def command(cmd, node_ip=None):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   执行shell命令
    :param cmd:     (str)要执行的命令
    :param node_ip: (str)节点ip,不输入时为本节点
    :return: 
    """
    if node_ip:
        cmd1 = 'ssh %s "%s"' % (node_ip, cmd)
    else:
        cmd1 = cmd
    process = subprocess.Popen(cmd1, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    output, unused_err = process.communicate()
    retcode = process.poll()
    return retcode, output


def run_command(cmd, fault_node_ip=None):
    if (fault_node_ip is not None) and (fault_node_ip in NODE_IP_LST):
        node_ips_list = NODE_IP_LST[:]
        node_ips_list.remove(fault_node_ip)
    else:
        node_ips_list = NODE_IP_LST[:]

    for node_ip in node_ips_list:
        # 判断节点是否可以ping通
        if check_ping(node_ip) is False:
            continue
        # 判断数据网是否正常
        if check_datanet(node_ip) is False:
            continue
        # 判断节点上是否有/home/parastor/conf
        if 0 != check_path(node_ip, '/home/parastor/conf'):
            continue
        # 判断节点上是否有集群
        rc, stdout = command(cmd, node_ip)
        if rc == 127:
            continue
        if (rc != 0) and ('FindMasterError' in stdout.strip().splitlines()[-1]):
            num = 1
            logging.warn('%s return "FindMasterError" %d times' % (cmd, num))
            while True:
                time.sleep(20)
                num += 1
                rc, stdout = command(cmd, node_ip)
                if (rc != 0) and ('FindMasterError' in stdout.strip().splitlines()[-1]):
                    logging.warn('%s return "FindMasterError" %d times' % (cmd, num))
                else:
                    break
        return rc, stdout


def json_loads(stdout):
    """
    :author:       baoruobing
    :date  :       2018.08.15
    :description:  json解析
    :param stdout: (str)要解析的字符串
    :return: 
    """
    try:
        stdout_str = json.loads(stdout, strict=False)
        return stdout_str
    except Exception, e:
        logging.error(stdout)
        raise Exception("Error msg is %s" % e)


def get_nodes_id():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取集群所有节点的id
    :return: 
    """
    cmd = "pscli --command=get_nodes"
    nodes_ids = []
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        node_info = json_loads(stdout)
        nodes = node_info['result']['nodes']
        for node in nodes:
            nodes_ids.append(node['node_id'])
    return nodes_ids


def get_nodes_ip():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取集群所有节点的ip
    :return: 
    """
    cmd = "pscli --command=get_nodes"
    nodes_ips = []
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        node_info = json_loads(stdout)
        nodes = node_info['result']['nodes']
        for node in nodes:
            nodes_ips.append(node['ctl_ips'][0]['ip_address'])
    return nodes_ips


def get_mgr_node_ids():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取所有管理节点
    :return: 
    """
    cmd = 'pscli --command=get_nodes'
    mgr_node_id_lst = []
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        msg = json_loads(stdout)
        nodes_info = msg["result"]["nodes"]
        for node_info in nodes_info:
            for service_info in node_info['services']:
                if service_info['service_type'] == 'oJmgs':
                    mgr_node_id_lst.append(node_info['node_id'])
                    break
    return mgr_node_id_lst


def get_node_id_by_ip(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   通过节点ip获取节点id
    :param node_ip: 节点ip
    :return: 
    """
    cmd = 'pscli --command=get_nodes'
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        msg = json_loads(stdout)
        nodes_info = msg["result"]["nodes"]
        for node in nodes_info:
            ctl_ip = node["ctl_ips"][0]["ip_address"]
            if node_ip == ctl_ip:
                return node["node_id"]
        logging.info("there is not a node's ip is %s!!!" % node_ip)
        return None


def get_node_ip_by_id(node_id):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   通过节点id获取节点ip
    :param node_id: (int)节点id
    :return: 
    """
    cmd = "pscli --command=get_nodes --ids=%s" % node_id
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        msg = json_loads(stdout)
        node_info = msg["result"]["nodes"][0]
        node_ip = node_info['ctl_ips'][0]['ip_address']
        return node_ip


def get_nodes_ips_by_ip(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   获取集群所有节点的管理ip
    :param node_ip: (str)节点ip
    :return: 
    """
    cmd = '"pscli --command=get_nodes"'
    rc, stdout = command(cmd, node_ip)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        node_ip_lst = []
        node_info = json_loads(stdout)
        nodes = node_info['result']['nodes']
        for node in nodes:
            node_ip_lst.append(node['ctl_ips'][0]['ip_address'])

    return node_ip_lst


def get_nodes_data_ip_by_ip(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   获取集群所有节点的数据ip
    :param node_ip: (str)节点ip
    :return: 
    """
    cmd = '"pscli --command=get_nodes"'
    rc, stdout = command(cmd, node_ip)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    data_ip_lst = []
    stdout = json_loads(stdout)
    node_info_lst = stdout['result']['nodes']
    for node_info in node_info_lst:
        for data_ip_info in node_info['data_ips']:
            data_ip_lst.append(data_ip_info['ip_address'])
    return data_ip_lst


def get_node_lnode_rel():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取node和lnode的关系
    :return:      (dic)键是nodeid, 值是lnodeid
    """
    """获取所有节点的id"""
    node_id_lst = get_nodes_id()
    cmd = "/home/parastor/tools/nWatch -t oRole -i %s -c oRole#rolemgr_view_dump" % node_id_lst[0]
    rc, stdout = run_command(cmd)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    line_lst = stdout.splitlines()
    i = 0
    for i in range(len(line_lst)):
        if "jtype:1 info" in line_lst[i]:
            break
    line_lst1 = line_lst[:i]
    node_lnode_rel = {}
    lnode_lst = []
    for line in line_lst1:
        if "-->--> node_sn:" in line:
            lnode_lst = []
            node_id = int(line.split(',')[1].split()[-1])
            node_lnode_rel[node_id] = lnode_lst
        if "-->-->--> lnodeid" in line:
            lnode_id = int(line.split(',')[0].split()[-1])
            lnode_lst.append(lnode_id)
    return node_lnode_rel


def get_lnode_by_path(mount_path):
    """
    :author:           baoruobing
    :date  :           2018.08.15
    :description:      获取某个目录的lnodeid
    :param mount_path: 挂载目录
    :return: 
    """
    cmd = "/home/parastor/tools/para_layinfo %s" % mount_path
    rc, stdout = run_command(cmd)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    line_lst = stdout.splitlines()
    lnode_id = 0
    for line in line_lst:
        if "lmosid" in line:
            lnode_id = int(line.split(':')[-1])
    return lnode_id


def get_mos_node_id():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取mos节点的id
    :return: 
    """
    """获取node和lnode的关系"""
    node_lnode_rel = get_node_lnode_rel()
    """获取挂载目录所属的lnode"""
    lnode_id = get_lnode_by_path(Anchor_Path)
    for node_id in node_lnode_rel:
        if lnode_id in node_lnode_rel[node_id]:
            return node_id
    return None


def get_mgr_nodes_id():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取所有管理节点的id
    :return: 
    """
    cmd = "pscli --command=get_services"
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        msg = json_loads(stdout)
        nodes_info_lst = msg['result']['nodes']
        mgr_node_id_lst = []
        for node_info in nodes_info_lst:
            node_id = node_info['node_id']
            services_lst = node_info['services']
            for service in services_lst:
                if service['service_type'] == 'oJmgs':
                    mgr_node_id_lst.append(node_id)
                    break
    return mgr_node_id_lst


def get_meta_data_disk_names(node_id):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   获取一个节点的元数据盘和数据盘
    :param node_id: (int)节点id
    :return: 
    """
    cmd = ("pscli --command=get_disks --node_ids=%s" % node_id)
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        msg = json_loads(stdout)
        meta_disk_names = []
        data_disk_names = []
        disks_pool = msg['result']['disks']
        for disk in disks_pool:
            if disk['usage'] == 'SHARED' and disk['usedState'] == 'IN_USE' and disk['state'] == 'DISK_STATE_HEALTHY':
                meta_disk_names.append(disk['devname'])
            elif disk['usage'] == 'DATA' and disk['usedState'] == 'IN_USE' and disk['state'] == 'DISK_STATE_HEALTHY':
                data_disk_names.append(disk['devname'])
    return meta_disk_names, data_disk_names


def get_scsiid_by_name(node_ip, disk_name):
    """
    :author:          baoruobing
    :date  :          2018.08.15
    :description:     获取一个磁盘的scsi_id
    :param node_ip:   节点ip
    :param disk_name: 磁盘名字
    :return:          磁盘scsi_id
    """
    cmd = 'lsscsi'
    rc, stdout = command(cmd, node_ip)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        list_stdout = stdout.strip().split('\n')
        for mem in list_stdout:
            if disk_name in mem:
                list_mem = mem.split()
                id = list_mem[0]
                id = id[1:-1]
                return id
    return None


def get_disk_usage_by_name(node_id, disk_name):
    """
    :author:          baoruobing
    :date  :          2018.08.15
    :description:     获取一个磁盘的用途
    :param node_id:   节点id
    :param disk_name: 磁盘name
    :return: 
    """
    cmd = "pscli --command=get_disks --node_ids=%s" % node_id
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        result = json_loads(stdout)
        disk_list = result['result']['disks']
        for disk in disk_list:
            if disk['devname'] == disk_name:
                return disk['usage']
    return None


def get_disk_uuid_by_name(node_id, disk_name):
    """
    :author:          baoruobing
    :date  :          2018.08.15
    :description:     获取一个磁盘的uuid
    :param node_id:   节点id
    :param disk_name: 磁盘name
    :return: 
    """
    cmd = "pscli --command=get_disks --node_ids=%s" % node_id
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        result = json_loads(stdout)
        disk_list = result['result']['disks']
        for disk in disk_list:
            if disk['devname'] == disk_name:
                return disk['uuid']
    return None


def get_disk_id_by_uuid(node_id, disk_uuid):
    """
    :author:          baoruobing
    :date  :          2018.08.15
    :description:     通过uuid获取一个磁盘的id
    :param node_id:   (int)节点id
    :param disk_uuid: (str)磁盘uuid
    :return: 
    """
    cmd = "pscli --command=get_disks --node_ids=%s" % node_id
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        result = json_loads(stdout)
        disk_list = result['result']['disks']
        for disk in disk_list:
            if disk['uuid'] == disk_uuid:
                return disk['id']
    return None


def get_storage_pool_id_by_diskid(node_id, disk_id):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   通过磁盘id获取所属的存储池的id
    :param node_id: (int)节点id
    :param disk_id: (id)磁盘id
    :return: 
    """
    cmd = 'pscli --command=get_disks --node_ids=%s' % node_id
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        msg = json_loads(stdout)
        disks_info = msg["result"]["disks"]
        for disk in disks_info:
            if disk['id'] == disk_id:
                return disk['storagePoolId']
        logging.info("there is not a disk's id is %s!!!" % disk_id)
        return None


def expand_disk_2_storage_pool(storage_pool_id, disk_id):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           将磁盘添加到存储池中
    :param storage_pool_id: 存储池id
    :param disk_id:         磁盘id
    :return: 
    """
    cmd = 'pscli --command=expand_storage_pool --storage_pool_id=%s --disk_ids=%s' % (storage_pool_id, disk_id)
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    return


def remove_disk(node_ip, disk_id, disk_usage):
    """
    :author:           baoruobing
    :date  :           2018.08.15
    :description:      拔盘
    :param node_ip:    节点ip
    :param disk_id:    磁盘id
    :param disk_usage: 磁盘用途
    :return: 
    """
    cmd = 'echo scsi remove-single-device %s > /proc/scsi/scsi' % disk_id
    logging.info('node %s remove disk %s, disk usage is %s' % (node_ip, disk_id, disk_usage))
    rc, stdout = command(cmd, node_ip)
    if 0 != rc:
        logging.error('node %s remove disk %s fault!!!' % (node_ip, disk_id))
    return


def insert_disk(node_ip, disk_id, disk_usage):
    """
    :author:           baoruobing
    :date  :           2018.08.15
    :description:      插盘
    :param node_ip:    节点ip
    :param disk_id:    磁盘id
    :param disk_usage: 磁盘用途
    :return: 
    """
    cmd = 'echo scsi add-single-device %s > /proc/scsi/scsi' % disk_id
    logging.info('node %s add disk %s, disk usage is %s' % (node_ip, disk_id, disk_usage))
    rc, stdout = command(cmd, node_ip)
    if 0 != rc:
        logging.error('node %s add disk %s fault!!!' % (node_ip, disk_id))
    time.sleep(5)
    cmd = 'lsscsi'
    rc, stdout = command(cmd, node_ip)
    logging.info(stdout)
    return


def delete_disk(disk_id, auto_query=True):
    """
    :author:           baoruobing
    :date  :           2018.08.15
    :description:      删盘
    :param disk_id:    (int)磁盘id
    :param auto_query: (bool)为True时,同步删除;为False时,异步删除. 默认:True
    :return: 
    """
    if auto_query:
        cmd = "pscli --command=remove_disks --disk_ids=%s" % disk_id
    else:
        cmd = "pscli --command=remove_disks --disk_ids=%s --auto_query=false" % disk_id
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    return


def add_disk(node_id, disk_uuid, disk_usage):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   添盘
    :param node_id: (int)节点id
    :param uuid:    (str)磁盘uuid
    :param usage:   (str)磁盘用途
    :return: 
    """
    cmd = ("pscli --command=add_disks --node_ids=%s --disk_uuids=%s --usage=%s" % (node_id, disk_uuid, disk_usage))
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    return


def check_disk_status(node_id, disk_uuid):
    """
    :author:          baoruobing
    :date  :          2018.08.15
    :description:     检查磁盘是否存在
    :param node_id:   节点id
    :param disk_uuid: 磁盘uuid
    :return: 
    """
    cmd = ("pscli --command=get_disks --node_ids=%s" % node_id)
    rc, stdout = run_command(cmd)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    stdout = json_loads(stdout)
    disk_info_lst = stdout['result']['disks']
    for disk_info in disk_info_lst:
        if disk_info['uuid'] == disk_uuid:
            if disk_info['state'] == 'DISK_STATE_HEALTHY':
                return True
            else:
                return False
    return False


def check_ip(ip):
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 检查ip的格式是否正确
    :param ip:    (str)要检查的ip
    :return: 
    """
    pattern = re.compile(r'((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    match = pattern.match(ip)
    if match:
        return True
    else:
        return False


def check_ping(ip):
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 检查ip是否可以ping通
    :param ip:    (str)检查的ip
    :return: 
    """
    cmd = 'ping -c 3 %s | grep "0 received" | wc -l' % ip
    rc, stdout = command(cmd)
    if '0' != stdout.strip():
        return False
    else:
        return True


def check_client_node(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   检查客户端状态
    :param node_ip: 
    :return: 
    """
    """检查ip是否可以ping通"""
    start_time = time.time()
    while True:
        if check_ping(node_ip):
            break
        exist_time = int(time.time() - start_time)
        m, s = divmod(exist_time, 60)
        h, m = divmod(m, 60)
        time_str = "node:%s can't ping %dh:%dm:%ds" % (node_ip, h, m, s)
        logging.info(time_str)
        time.sleep(10)

    """检查客户端挂载路径是否正常"""
    cmd = "df"



def get_data_net_eth(node_id, node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   获取一个节点所有数据网的网卡名字、ip和掩码
    :param node_id: (int)节点id
    :param node_ip: (str)节点ip
    :return:        (list)成员是字典{"eth":"eth1", "ip":"10.2.41.101", "mask":"255.255.252.0"}
    """
    '''获取本节点的数据网的eth名字'''
    cmd = "pscli --command=get_nodes --ids=%s" % node_id

    data_ip_list = []
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        result = json_loads(stdout)
        data_ips = result['result']['nodes'][0]['data_ips']
        for data_ip in data_ips:
            ip = data_ip['ip_address']
            data_ip_list.append(ip)

    eth_list = []
    for ip in data_ip_list:
        tem_dic = {}
        cmd1 = 'ip addr | grep %s' % ip
        rc, stdout = command(cmd1, node_ip)
        if 0 != rc:
            logging.warn("ip %s is not exist" % ip)
            continue
        else:
            eth_name = stdout.split()[-1]
            tem_dic["eth"] = eth_name

        cmd2 = 'ifconfig | grep %s' % ip
        rc, stdout = command(cmd2, node_ip)
        if 0 != rc:
            raise Exception(
                "Execute command: \"%s\" failed. \nstdout: %s" % (cmd2, stdout))
        else:
            mask = stdout.strip().split()[3]
            tem_dic["dataip"] = ip
            tem_dic["mgrip"] = node_ip
            tem_dic["mask"] = mask

        eth_list.append(tem_dic)
    return eth_list


def get_net_eth(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   获取一个节点的所有网口
    :param node_ip: 节点ip 
    :return: 
    """
    cmd = "ip addr | grep \\\"scope global\\\""
    rc, stdout = command(cmd, node_ip)
    if rc != 0:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    line_lst = stdout.splitlines()
    ip_lst = []
    for line in line_lst:
        ip_lst.append(line.split()[1].split('/')[0])
    eth_list = []
    for ip in ip_lst:
        tem_dic = {}
        cmd1 = 'ip addr | grep %s' % ip
        rc, stdout = command(cmd1, node_ip)
        if 0 != rc:
            raise Exception(
                "Execute command: \"%s\" failed. \nstdout: %s" % (cmd1, stdout))
        else:
            eth_name = stdout.split()[-1]
            tem_dic["eth"] = eth_name

        cmd2 = 'ifconfig | grep %s' % ip
        rc, stdout = command(cmd2, node_ip)
        if 0 != rc:
            raise Exception(
                "Execute command: \"%s\" failed. \nstdout: %s" % (cmd2, stdout))
        else:
            mask = stdout.strip().split()[3]
            tem_dic["dataip"] = ip
            tem_dic["mask"] = mask
        eth_list.append(tem_dic)
    return eth_list


def get_client_data_net(client_ip):
    """
    :author:          baoruobing
    :date  :          2018.08.15
    :description:     获取一个客户端的数据网网口
    :param client_ip: 客户端ip
    :return: 
    """
    """随机获取一个集群节点"""
    node_id_lst = get_nodes_id()
    node_id = random.choice(node_id_lst)
    node_ip = get_node_ip_by_id(node_id)
    date_eth_lst = get_data_net_eth(node_id, node_ip)

    client_eth_lst = get_net_eth(client_ip)
    client_data_eth_lst = []
    for data_eth in date_eth_lst:
        for client_eth in client_eth_lst:
            if check_ip_route(data_eth['dataip'], data_eth['mask'], client_eth['dataip'], data_eth['mask']):
                client_data_eth_lst.append(client_eth)
    return client_data_eth_lst


def check_ip_route(ip1, mask1, ip2, mask2):
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 检查两个ip的路由是否相同
    :param ip1:   
    :param mask1: 
    :param ip2: 
    :param mask2: 
    :return: 
    """
    ip1_lst = ip1.split('.')
    mask1_lst = mask1.split('.')
    ip2_lst = ip2.split('.')
    mask2_lst = mask2.split('.')

    route1_lst = []
    route2_lst = []
    for i in range(4):
        route1_lst.append(str(int(ip1_lst[i]) & int(mask1_lst[i])))
        route2_lst.append(str(int(ip2_lst[i]) & int(mask2_lst[i])))

    route1 = '.'.join(route1_lst)
    route2 = '.'.join(route2_lst)

    if route1 == route2:
        return True
    else:
        return False
    

def run_down_net(node_ip, eth_lst):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   down节点的网卡
    :param node_ip: (str)节点ip
    :param eth_lst: (list)网卡的列表
    :return: 
    """
    for eth in eth_lst:
        cmd = 'ifdown %s' % eth
        logging.info("node %s ifdown %s" % (node_ip, eth))
        rc, stdout = command(cmd, node_ip)
        if 0 != rc:
            logging.warn("node %s  ifdown %s failed!!!" % (node_ip, eth))
    return


def run_up_net(node_ip, eth_lst):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   up节点的网卡
    :param node_ip: (str)节点ip
    :param eth_lst: (list)网卡的列表
    :return: 
    """
    for eth in eth_lst:
        cmd = 'ifup %s' % eth
        logging.info("node %s ifup %s" % (node_ip, eth))
        rc, stdout = command(cmd, node_ip)
        if 0 != rc:
            logging.warn("node %s  ifup %s failed!!!" % (node_ip, eth))
    return


def check_datanet(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   检查节点的数据网是否存在
    :param node_ip: 节点ip
    :return: 
    """
    cmd = '"ip addr | grep "inet ""'
    rc, stdout = command(cmd, node_ip)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    lines = stdout.strip().split('\n')
    for line in lines:
        ip = line.split()[1].split('/')[0]
        if ip in NODE_DATA_IP_LST:
            return True
    return False


def check_process(node_ip, process):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   检查节点某个进程是否存在
    :param node_ip: 节点ip
    :param process: 进程名字
    :return: 
    """
    ps_cmd = ('ps -ef | grep %s | grep -v grep' % process)
    rc, stdout = command(ps_cmd, node_ip)
    if 0 == rc:
        return True
    else:
        return False


def run_kill_process(node_ip, process):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   kill节点某个进程
    :param node_ip: 节点ip
    :param process: 进程名字
    :return: 
    """
    ps_cmd = ('ps -ef | grep %s | grep -v grep' % process)
    rc, stdout = command(ps_cmd, node_ip)
    if '' == stdout:
        return
    logging.info(stdout)
    lines = stdout.strip().split('\n')
    for line in lines:
        vars = line.split()
        pid = vars[1]
        kill_cmd = ('kill -9 %s' % pid)
        logging.info('node %s kill %s' % (node_ip, process))
        rc, stdout = command(kill_cmd, node_ip)
        if 0 != rc:
            logging.error("Execute command: \"%s\" failed. \nstdout: %s" % (kill_cmd, stdout))
    return


def kill_process(node_ip, process):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   kill进程故障
    :param node_ip: 节点ip
    :param process: 进程名字
    :return: 
    """
    """kill pro"""
    run_kill_process(node_ip, process)

    """不断检查进程是否起来"""
    start_time = time.time()
    while True:
        logging.info("wait 20 s")
        time.sleep(20)
        if check_process(node_ip, process) is False:
            exist_time = int(time.time() - start_time)
            m, s = divmod(exist_time, 60)
            h, m = divmod(m, 60)
            logging.info('node:%s, process:%s is not OK %dh:%dm:%ds!!!' % (node_ip, process, h, m, s))
        else:
            break
    logging.info("node:%s, process:%s is OK" % (node_ip, process))
    return


def check_node_master_zk(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   检查节点是否是主zk节点
    :param node_ip: 节点ip
    :return: 
    """
    cmd = "sh /root/zk/bin/zkServer.sh status"
    rc, stdout = command(cmd, node_ip)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    line_lst = stdout.splitlines()
    zk_type = line_lst[-1].split()[-1]
    if zk_type == "leader":
        return True
    else:
        return False


def get_master_orole_node_id():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获得主oRole节点的id
    :return: 
    """
    """获取所有节点的id"""
    node_id_lst = get_nodes_id()
    """获取主oRole节点"""
    cmd = '/home/parastor/tools/nWatch -t oRole -i %s -c oRole#rolemgr_master_dump' % node_id_lst[0]
    rc, stdout = run_command(cmd)
    if rc != 0 or 'failed' in stdout.strip().splitlines()[0]:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    master_orole_id = int(stdout.split(':')[-1].strip())
    return master_orole_id


def get_master_ojob_node_id():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获得主ojob节点的id
    :return: 
    """
    """获取所有节点的id"""
    node_id_lst = get_nodes_id()
    """获取主oRole节点"""
    cmd = '/home/parastor/tools/nWatch -t oJob -i %s -c RCVR#jobinfo' % node_id_lst[0]
    rc, stdout = run_command(cmd)
    if rc != 0 or 'failed' in stdout.strip().splitlines()[0]:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    master_job_id = int(stdout.split(',')[0].split()[-1])
    return master_job_id


def check_path(node_ip, path):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   检查节点上的某个路径是否存在
    :param node_ip: 节点ip
    :param path:    要检查的路径
    :return: 
    """
    cmd = 'ls %s' % path
    rc, stdout = command(cmd, node_ip)
    return rc


def get_all_volume_layout():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取所有卷的配比
    :return:      (list)卷的配比信息,[{'disk_parity_num':2,'node_parity_num':1,'replica_num':4}]
    """
    cmd = "pscli --command=get_volumes"
    rc, stdout = run_command(cmd)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    volumes_info = json_loads(stdout)
    volumes_lst = volumes_info['result']['volumes']
    layout_lst = []
    for volume in volumes_lst:
        layout_dic = {}
        layout_dic['disk_parity_num'] = volume['layout']['disk_parity_num']
        layout_dic['node_parity_num'] = volume['layout']['node_parity_num']
        layout_dic['replica_num'] = volume['layout']['replica_num']
        layout_lst.append(layout_dic)
    return layout_lst


def check_case_layout(nodenum, node_parity_num, disk_parity_num):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           检查节点上的某个路径是否存在
    :param nodenum:         (int)用例的节点个数
    :param node_parity_num: (int)用例的节点冗余数
    :param disk_parity_num: (int)用例的磁盘冗余数
    :return: 
    """
    """获取集群节点的个数"""
    node_id_lst = get_nodes_id()
    para_node_num = len(node_id_lst)

    """判断节点个数是否满足"""
    if nodenum > para_node_num:
        logging.warn("total num of node not meet the requirements!")
        return False

    """获取所有卷的配比信息"""
    layout_lst = get_all_volume_layout()

    """获取最小的磁盘冗余和节点冗余"""
    node_parity_num_tmp = 1000
    disk_parity_num_tmp = 1000
    for layout in layout_lst:
        if layout['disk_parity_num'] != 0:
            disk_parity_num_tmp = layout['disk_parity_num'] < disk_parity_num_tmp and layout[
                'disk_parity_num'] or disk_parity_num_tmp
        else:
            disk_parity_num_tmp = layout['replica_num'] - 1 < disk_parity_num_tmp and layout[
                'replica_num'] - 1 or disk_parity_num_tmp
        node_parity_num_tmp = layout['node_parity_num'] < node_parity_num_tmp and layout[
            'node_parity_num'] or node_parity_num_tmp

    """判断节点冗余是否满足"""
    if node_parity_num > node_parity_num_tmp:
        logging.warn("node parity num not meet the requirements!")
        return False

    """判断磁盘冗余是否满足"""
    if disk_parity_num > disk_parity_num_tmp:
        logging.warn("disk parity num not meet the requirements!")
        return False
    return True


def get_sys_ip():
    """
    :author:     baoruobing
    :date  :     2018.08.15
    :description:获取所有节点包括客户端的节点ip
    :return: 
    """
    sys_ip_lst = []
    cmd = "pscli --command=get_clients"
    rc, result = run_command(cmd)
    if 0 != rc:
        raise Exception("There is not parastor or get nodes ip failed!!!")
    else:
        result = json_loads(result)
        nodes_lst = result['result']
        for node in nodes_lst:
            sys_ip_lst.append(node['ip'])
    return sys_ip_lst


def wait_time(fault):
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 根据故障类型等待多长时间
    :param fault: (str)故障类型,wait_times的键
    :return: 
    """
    time_lst = wait_times.get(fault)
    min_time = time_lst[0]
    max_time = time_lst[1]

    wait_time = random.randint(min_time, max_time)
    logging.info("wait %d s" % wait_time)
    time.sleep(wait_time)
    return

"""****************************** 故障函数 ******************************"""


def get_master_mgr_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取主管理节点的id
    :return:      (str)主管理节点的管理ip
    """
    cmd = "pscli --command=get_master"
    rc, stdout = run_command(cmd)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    result = json_loads(stdout)
    node_id = result['result']['node_id']
    cmd = "pscli --command=get_nodes --ids=%s" % node_id
    rc, stdout = run_command(cmd)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    result = json_loads(stdout)
    node_ip = result['result']['nodes'][0]['ctl_ips'][0]['ip_address']
    return node_ip, None


def get_no_master_mgr_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取一个非管理主节点
    :return: 
    """
    """获取非管理节点id"""
    master_node_ip, disk_name = get_master_mgr_node()
    """获取所有节点的id"""
    nodes_ip_lst = get_nodes_ip()
    while True:
        fault_node_ip = random.choice(nodes_ip_lst)
        if fault_node_ip != master_node_ip:
            break
    return fault_node_ip, None


def get_mgr_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取一个管理节点
    :return: 
    """
    mgr_node_id_lst = get_mgr_nodes_id()
    mgr_node_id = random.choice(mgr_node_id_lst)
    mgr_node_ip = get_node_ip_by_id(mgr_node_id)
    return mgr_node_ip, None


def get_no_mgr_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取一个非管理节点
    :return: 
    """
    """获取管理节点id"""
    mgr_node_id_lst = get_mgr_nodes_id()
    """获取所有节点的id"""
    nodes_id_lst = get_nodes_id()
    no_mgr_node_id_lst = list(set(mgr_node_id_lst) ^ set(nodes_id_lst))
    if no_mgr_node_id_lst:
        no_mgr_node_id = random.choice(no_mgr_node_id_lst)
        no_mgr_node_ip = get_node_ip_by_id(no_mgr_node_id)
        return no_mgr_node_ip, None
    else:
        return None, None


def get_one_client_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取一个独立私有客户端节点
    :return: 
    """
    cmd = "pscli --command=get_clients"
    rc, stdout = run_command(cmd)
    if rc != 0:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    result = json_loads(stdout)
    ext_client_ip_lst = []
    client_info_lst = result['result']
    for client_info in client_info_lst:
        if client_info['type'] == 'EXTERNAL':
            ext_client_ip_lst.append(client_info['ip'])
    client_ip = random.choice(ext_client_ip_lst)
    return client_ip, None


def get_lmos_node_and_meta_disk():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取lmos节点和节点上的一个元数据盘
    :return: 
    """
    """获取lmos节点"""
    lmos_node_id = get_mos_node_id()
    """获取节点的所有数据盘和共享盘"""
    meta_disk_name_lst, data_disk_name_lst = get_meta_data_disk_names(lmos_node_id)
    """随机获取一个元数据盘"""
    meta_disk_name = random.choice(meta_disk_name_lst)
    """获取lmos节点的ip"""
    lmos_node_ip = get_node_ip_by_id(lmos_node_id)
    return lmos_node_ip, meta_disk_name


def get_no_lmos_node_and_meta_disk():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取非lmos节点和节点上的一个元数据盘
    :return: 
    """
    """获取lmos节点"""
    lmos_node_id = get_mos_node_id()
    """随机获取非lmos节点id"""
    node_id_lst = get_nodes_id()
    fault_node_id = 0
    for node_id in node_id_lst:
        if node_id != lmos_node_id:
            fault_node_id = node_id
            break

    """获取节点的所有数据盘和共享盘"""
    meta_disk_name_lst, data_disk_name_lst = get_meta_data_disk_names(fault_node_id)
    """随机获取一个元数据盘"""
    meta_disk_name = random.choice(meta_disk_name_lst)
    """获取lmos节点的ip"""
    fault_node_ip = get_node_ip_by_id(fault_node_id)
    return fault_node_ip, meta_disk_name


def get_random_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 随机获取一个节点
    :return: 
    """
    """获取所有节点的ip"""
    node_ip_lst = get_nodes_ip()
    """随机获取一个节点"""
    fault_node_ip = random.choice(node_ip_lst)
    return fault_node_ip, None


def get_random_node_data_disk():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 随机获取一个节点和一个数据盘
    :return: 
    """
    """获取所有节点的id"""
    node_id_lst = get_nodes_id()
    """随机获取一个节点"""
    fault_node_id = random.choice(node_id_lst)

    """获取节点的所有数据盘和共享盘"""
    meta_disk_name_lst, data_disk_name_lst = get_meta_data_disk_names(fault_node_id)
    """随机获取一个数据盘"""
    data_disk_name = random.choice(data_disk_name_lst)
    """获取故障节点的ip"""
    fault_node_ip = get_node_ip_by_id(fault_node_id)
    return fault_node_ip, data_disk_name


def get_master_zk_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取主zk节点
    :return: 
    """
    """获取主zk节点"""
    mgr_node_id_lst = get_mgr_node_ids()
    master_zk_node_ip = ''
    for node_id in mgr_node_id_lst:
        node_ip = get_node_ip_by_id(node_id)
        if check_node_master_zk(node_ip):
            master_zk_node_ip = node_ip
            break
    if master_zk_node_ip == '':
        raise Exception("there is no master zk node")
    return master_zk_node_ip, None


def get_lmos_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取lmos节点
    :return: 
    """
    """获取lmos节点"""
    lmos_node_id = get_mos_node_id()
    lmos_node_ip = get_node_ip_by_id(lmos_node_id)
    return lmos_node_ip, None


def get_no_lmos_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取非lmos节点
    :return: 
    """
    """获取lmos节点"""
    lmos_node_id = get_mos_node_id()
    """获取所有节点的id"""
    node_id_lst = get_nodes_id()
    while True:
        node_id = random.choice(node_id_lst)
        if node_id != lmos_node_id:
            fault_node_id = node_id
            break
    fault_node_ip = get_node_ip_by_id(fault_node_id)
    return fault_node_ip, None


def get_master_orole_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取主orole节点
    :return: 
    """
    orole_node_id = get_master_orole_node_id()
    orole_node_ip = get_node_ip_by_id(orole_node_id)
    return orole_node_ip, None


def get_no_master_orole_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取非主orole节点
    :return: 
    """
    """获取主oRole节点的id"""
    orole_node_id = get_master_orole_node_id()
    """获取所有节点的id"""
    node_id_lst = get_nodes_id()
    while True:
        node_id = random.choice(node_id_lst)
        if node_id != orole_node_id:
            fault_node_id = node_id
            break
    fault_node_ip = get_node_ip_by_id(fault_node_id)
    return fault_node_ip, None


def get_master_ojob_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取主ojob节点
    :return: 
    """
    master_ojob_id = get_master_ojob_node_id()
    master_ojob_ip = get_node_ip_by_id(master_ojob_id)
    return master_ojob_ip, None


def get_no_master_ojob_node():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 获取非主ojob节点
    :return: 
    """
    """获取主oJob节点的id"""
    master_ojob_id = get_master_ojob_node_id()
    """获取所有节点的id"""
    node_id_lst = get_nodes_id()
    while True:
        node_id = random.choice(node_id_lst)
        if node_id != master_ojob_id:
            fault_node_id = node_id
            break
    fault_node_ip = get_node_ip_by_id(fault_node_id)
    return fault_node_ip, None


def fault_all_net(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           执行一个点上全部网卡故障
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    """获取故障节点的ip"""
    fault_node_id = get_node_id_by_ip(fault_node_ip)
    """获取节点的所有数据网卡"""
    eth_info_lst = get_data_net_eth(fault_node_id, fault_node_ip)
    eth_lst = []
    for tem in eth_info_lst:
        eth_lst.append(tem['eth'])

    logging.info("node:%s, fault all net:%s" % (fault_node_ip, eth_lst))

    """down所有数据网"""
    run_down_net(fault_node_ip, eth_lst)

    wait_time('down_net')

    """up所有数据网"""
    run_up_net(fault_node_ip, eth_lst)

    time.sleep(30)
    return


def fault_half_net(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           执行一个点上一半网卡故障
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    """获取故障节点的ip"""
    fault_node_id = get_node_id_by_ip(fault_node_ip)
    """获取节点的所有数据网卡"""
    eth_info_lst = get_data_net_eth(fault_node_id, fault_node_ip)
    eth_lst = []
    for tem in eth_info_lst:
        eth_lst.append(tem['eth'])

    fault_eth_lst = random.sample(eth_lst, len(eth_lst)/2)
    if len(fault_eth_lst) == 0:
        logging.warn("node %s data eth is %s; can't down half net" % (fault_node_ip, eth_info_lst))
        return

    logging.info("node:%s, fault half net:%s" % (fault_node_ip, fault_eth_lst))

    """down数据网"""
    run_down_net(fault_node_ip, fault_eth_lst)

    wait_time('down_net')

    """up数据网"""
    run_up_net(fault_node_ip, fault_eth_lst)

    time.sleep(30)
    return


def fault_client_all_net(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           故障一个客户端上的所有数据网
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    """获取客户端上的所有数据网"""
    client_data_eth_lst = get_client_data_net(fault_node_ip)
    eth_lst = []
    for tem in client_data_eth_lst:
        eth_lst.append(tem['eth'])

    logging.info("client node:%s, fault all net:%s" % (fault_node_ip, eth_lst))

    """down所有数据网"""
    run_down_net(fault_node_ip, eth_lst)

    wait_time('down_net')

    """up所有数据网"""
    run_up_net(fault_node_ip, eth_lst)

    time.sleep(30)
    return


def fault_client_half_net(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           故障一个客户端上的一半数据网
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    """获取客户端上的所有数据网"""
    client_data_eth_lst = get_client_data_net(fault_node_ip)
    eth_lst = []
    for tem in client_data_eth_lst:
        eth_lst.append(tem['eth'])

    fault_eth_lst = random.sample(eth_lst, len(eth_lst) / 2)
    if len(fault_eth_lst) == 0:
        logging.warn("node %s data eth is %s; can't down half net" % (fault_node_ip, client_data_eth_lst))
        return

    logging.info("client node:%s, fault half net:%s" % (fault_node_ip, fault_eth_lst))

    """down所有数据网"""
    run_down_net(fault_node_ip, fault_eth_lst)

    wait_time('down_net')

    """up所有数据网"""
    run_up_net(fault_node_ip, fault_eth_lst)

    time.sleep(30)
    return


def fault_disk_no_rebuild(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           执行拔盘故障(不重建)
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: (str)故障盘名字
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    fault_disk_scsi_id = get_scsiid_by_name(fault_node_ip, fault_disk_name)
    fault_node_id = get_node_id_by_ip(fault_node_ip)
    fault_disk_usage = get_disk_usage_by_name(fault_node_id, fault_disk_name)
    fault_disk_uuid = get_disk_uuid_by_name(fault_node_id, fault_disk_name)

    logging.info("node:%s, pullout disk:%s  no rebuild" % (fault_node_ip, fault_disk_name))
    """拔盘"""
    remove_disk(fault_node_ip, fault_disk_scsi_id, fault_disk_usage)

    """等一段时间"""
    wait_time('down_disk')

    """插盘"""
    insert_disk(fault_node_ip, fault_disk_scsi_id, fault_disk_usage)

    """删除共享盘"""
    if fault_disk_usage == 'SHARED':
        logging.info(
            'delete node %s disk %s, disk usage is %s' % (fault_node_ip, fault_disk_name, fault_disk_usage))
        fault_disk_id_old = get_disk_id_by_uuid(fault_node_id, fault_disk_uuid)
        delete_disk(fault_disk_id_old)

    """检查磁盘是否可以查到"""
    start_time = int(time.time())
    while True:
        if check_disk_status(fault_node_id, fault_disk_uuid):
            logging.info("node:%s disk:%s exist" % (fault_node_ip, fault_disk_name))
            break
        else:
            now_time = int(time.time())
            interval_time = now_time - start_time
            logging.info("%ss, node:%s disk:%s is not exist!!!" % (interval_time, fault_node_ip, fault_disk_name))
            if interval_time > 300:
                logging.error("%ss, node:%s disk:%s is still not exist. ERROR!!!"
                              % (interval_time, fault_node_ip, fault_disk_name))
                raise Exception("%ss, node:%s disk:%s is still not exist. ERROR!!!"
                                % (interval_time, fault_node_ip, fault_disk_name))
            time.sleep(5)

    """添加共享盘"""
    if fault_disk_usage == 'SHARED':
        logging.info(
            'add node %s disk %s, disk usage is %s' % (fault_node_ip, fault_disk_name, fault_disk_usage))
        add_disk(fault_node_id, fault_disk_uuid, fault_disk_usage)


def fault_disk_rebuild(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           执行拔盘故障(重建)
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: (str)故障磁盘name
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    """修改磁盘超时参数"""
    cmd = "pscli --command=update_param --section=MGR --name=disk_isolate2rebuild_timeout --current=30000"
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))

    fault_disk_scsi_id = get_scsiid_by_name(fault_node_ip, fault_disk_name)
    fault_node_id = get_node_id_by_ip(fault_node_ip)
    fault_disk_uuid = get_disk_uuid_by_name(fault_node_id, fault_disk_name)
    fault_disk_usage = get_disk_usage_by_name(fault_node_id, fault_disk_name)

    logging.info("node:%s, pullout disk:%s  rebuild" % (fault_node_ip, fault_disk_name))

    """拔盘"""
    remove_disk(fault_node_ip, fault_disk_scsi_id, fault_disk_usage)

    logging.info("waiting 60s")
    time.sleep(60)

    """检查重建任务是否存在"""
    start_time = time.time()
    while True:
        if check_rebuild_job() is False:
            logging.info('rebuild job finish!!!')
            break
        time.sleep(20)
        exist_time = int(time.time() - start_time)
        m, s = divmod(exist_time, 60)
        h, m = divmod(m, 60)
        logging.info('rebuild job exist %dh:%dm:%ds' % (h, m, s))

    """检查坏对象"""
    check_badobj()

    """插盘"""
    insert_disk(fault_node_ip, fault_disk_scsi_id, fault_disk_usage)

    time.sleep(60)

    """删除磁盘"""
    fault_disk_id_old = get_disk_id_by_uuid(fault_node_id, fault_disk_uuid)
    storage_pool_id = get_storage_pool_id_by_diskid(fault_node_id, fault_disk_id_old)

    logging.info('delete node %s disk %s, disk usage is %s' % (fault_node_ip, fault_disk_name, fault_disk_usage))
    delete_disk(fault_disk_id_old)

    """检查磁盘是否可以查到"""
    start_time = int(time.time())
    while True:
        if check_disk_status(fault_node_id, fault_disk_uuid):
            logging.info("node:%s disk:%s exist" % (fault_node_ip, fault_disk_name))
            break
        else:
            now_time = int(time.time())
            interval_time = now_time - start_time
            logging.info("%ss, node:%s disk:%s is not exist!!!" % (interval_time, fault_node_ip, fault_disk_name))
            if interval_time > 300:
                logging.error("%ss, node:%s disk:%s is still not exist. ERROR!!!"
                              % (interval_time, fault_node_ip, fault_disk_name))
                raise Exception("%ss, node:%s disk:%s is still not exist. ERROR!!!"
                                % (interval_time, fault_node_ip, fault_disk_name))
            time.sleep(5)

    """加入磁盘"""
    logging.info('add node %s disk %s, disk usage is %s' % (fault_node_ip, fault_disk_name, fault_disk_usage))
    add_disk(fault_node_id, fault_disk_uuid, fault_disk_usage)

    """加入存储池"""
    if 'SHARED' != fault_disk_usage:
        fault_disk_id_new = get_disk_id_by_uuid(fault_node_id, fault_disk_uuid)
        logging.info('add node %s disk %s to storage_pool %s' % (fault_node_ip, fault_disk_name, storage_pool_id))
        expand_disk_2_storage_pool(storage_pool_id, fault_disk_id_new)

    """恢复磁盘超时参数"""
    cmd = "pscli --command=update_param --section=MGR --name=disk_isolate2rebuild_timeout --current=18000000"
    rc, stdout = run_command(cmd)
    if 0 != rc:
        raise Exception("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    return


def fault_master_zk_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill节点的zk进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("node:%s, kill process:zk" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['zk'])
    return


def fault_orole_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill节点的oRole进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("node:%s, kill process:oRole" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['oRole'])
    return


def fault_ostor_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill节点的oStor进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("node:%s, kill process:oStor" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['oStor'])
    return


def fault_ojob_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill节点的ojob进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("node:%s, kill process:oJob" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['oJob'])
    return


def fault_ojmgs_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill节点的ojmgs进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("node:%s, kill process:oJmgs" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['oJmgs'])
    return


def fault_omgcd_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill集群节点的omgcd进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("node:%s, kill process:oMgcd" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['oMgcd'])
    return


def fault_omgcd_client_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill客户端节点的omgcd进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("client node:%s, kill process:oMgcd_client" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['oMgcd_client'])
    return


def fault_opara_pro(fault_node_ip, fault_disk_name):
    """
    :author:                baoruobing
    :date  :                2018.08.15
    :description:           kill节点的opara进程
    :param fault_node_ip:   (str)故障节点ip
    :param fault_disk_name: None
    :return: 
    """
    """检查故障节点是否存在"""
    if not fault_node_ip:
        logging.warn("fault_node_ip is %s, %s skip" % (fault_node_ip, sys._getframe().f_code.co_name))
        return

    logging.info("node:%s, kill process:oPara" % fault_node_ip)
    kill_process(fault_node_ip, Process_Dic['oPara'])
    return

"""****************************** 环境检查 ******************************"""


def run_check_core(node_ip):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   检查core是否存在
    :param node_ip: 节点ip
    :return: 
    """
    core_path_lst = ['/home/parastor/log/', '/']
    for core_path in core_path_lst:
        core_path_tmp = os.path.join(core_path, 'core*')
        cmd = 'ls %s' % core_path_tmp
        rc, result = command(cmd, node_ip)
        if 0 != rc:
            return True
        else:
            return False


def check_core():
    """
    :author:     baoruobing
    :date  :     2018.08.15
    :description:检查集群中是否有core存在
    :return: 
    """
    flag = True
    core_node_lst = []
    sys_ip_lst = get_sys_ip()
    for node_ip in sys_ip_lst:
        # 先检查是否可以ping通
        if check_ping(node_ip) is False:
            logging.warn('node %s ping failed!!!' % node_ip)
            continue
        else:
            if run_check_core(node_ip) is False:
                flag = False
                core_node_lst.append(node_ip)
    if flag is False:
        core_node = ','.join(core_node_lst)
        logging.warn("These nodes %s has core!!! ", core_node)
        if CORESTOP is True:
            sys.exit(-1)
    else:
        logging.info("The current environment does not have core")
    return


def check_rebuild_job(fault_node_ip=None):
    """
    :author:              baoruobing
    :date  :              2018.08.15
    :description:         检查集群中是否有core存在
    :param fault_node_ip: 故障节点ip,在网络故障时使用,避免pscli执行在网络故障节点上
    :return: 
    """
    cmd = 'pscli --command=get_jobengine_state'
    rc, stdout = run_command(cmd, fault_node_ip)
    if 0 != rc:
        raise Exception(
            "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
    else:
        msg = json_loads(stdout)
        jobs_info = msg["result"]["job_engines"]
        for job in jobs_info:
            if job['type'] == 'JOB_ENGINE_REBUILD':
                return True
        return False


def run_check_badobj(node_id):
    """
    #todo 用oPara还是oJob
    node_id_lst = get_nodes_id()
    badobj_num = 0
    for node_id in node_id_lst:
        cmd = "/home/parastor/tools/nWatch -t oPara -i %d -c RCVR#badobj" % node_id
        rc, stdout = run_command(cmd)
        if 0 != rc:
            logging.error("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
            raise Exception(
                "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
        badobj_num += int(stdout.splitlines()[1].strip().split()[-1].strip())

    if badobj_num != 0:
        logging.info("badobj_num = %d" % (badobj_num))
        return False
    """
    cmd = "/home/parastor/tools/nWatch -t oJob -i %s -c RCVR#jobinfo" % node_id
    rc, stdout = run_command(cmd)
    if 0 != rc or 'failed' in stdout.strip().splitlines()[0]:
        logging.warn("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
        return -1
    master_job_id = stdout.split(',')[0].split()[-1]

    cmd = "/home/parastor/tools/nWatch -t oJob -i %s -c RCVR#repairjob" % master_job_id
    rc, result_badobj = run_command(cmd)
    if 0 != rc or 'failed' in stdout.strip().splitlines()[0]:
        logging.warn("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, result_badobj))
        return -1
    result_tmp = result_badobj.split()
    if "0" != result_tmp[-3]:
        logging.info("masterjob = %s, badobj_num = %s" % (master_job_id, result_tmp[-3]))
        return 1

    logging.info("The current environment does not have badobj")
    return 0


def check_badobj(waitflag=True, fault_ip=None):
    """
    :author:         baoruobing
    :date  :         2018.08.15
    :description:    每隔一段时间检查一遍是否还有坏对象
    :param waitflag: (bool)一开始是否需要等待一段时间
    :param fault_ip: (str)故障节点ip(断网的时候填入)
    :return: 
    """
    if waitflag is True:
        # 等待60s
        logging.info("wait %ds" % CHECKBADOBJ_WAIT_TIME)
        time.sleep(CHECKBADOBJ_WAIT_TIME)

    def _check_badjob():
        node_ip_lst_bak = NODE_IP_LST[:]
        if fault_ip:
            if fault_ip in node_ip_lst_bak:
                node_ip_lst_bak.remove(fault_ip)
        for node_ip in node_ip_lst_bak:
            # 检查是否可以ping通
            if check_ping(node_ip) is False:
                continue
            node_id = get_node_id_by_ip(node_ip)
            result = run_check_badobj(node_id)
            if -1 == result:
                continue
            elif 1 == result:
                return 1
            else:
                return 0

    start_time = time.time()
    while True:
        time.sleep(20)
        if 0 == _check_badjob():
            break
        exist_time = int(time.time() - start_time)
        m, s = divmod(exist_time, 60)
        h, m = divmod(m, 60)
        time_str = "badobj exist %dh:%dm:%ds" % (h, m, s)
        logging.info(time_str)
    return


def run_check_vset():
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   检查是否还有vset没有展平
    :return: 
    """
    node_ids = get_nodes_id()
    node_id = random.choice(node_ids)
    cmd = '/home/parastor/tools/nWatch -i %s -t oRole -c oRole#rolemgr_view_dump' % node_id
    rc, stdout = run_command(cmd)
    if 0 != rc or 'failed' in stdout.strip().splitlines()[0]:
        logging.warn("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
        return False
    stdout_lst = stdout.strip().splitlines()
    index = None
    for line in stdout_lst:
        if 'jtype:1 info' in line:
            index = stdout_lst.index(line)
            break
    if index is None:
        logging.warn("get mgrid failed!!!")
        return False
    stdout_lst1 = stdout_lst[index + 1:]

    for line in stdout_lst1:
        if 'jtype:' in line:
            index = stdout_lst1.index(line)
            stdout_lst2 = stdout_lst1[:index]
            break
    else:
        stdout_lst2 = stdout_lst1[:]

    index = None
    for line in stdout_lst2:
        if 'grpview info' in line:
            index = stdout_lst2.index(line)
    if index is None:
        logging.warn("get mgrid failed!!!")
        return False
    stdout_grpview_lst = stdout_lst2[index:]
    """支持多vmgr的场景"""
    node_lnode_lst = []
    lnode_dic = {}
    lnodeid_lst = []
    for line in stdout_grpview_lst:
        if '-->-->--> lnodeid:' not in line:
            if lnode_dic and 'lnode_id' in lnode_dic:
                node_lnode_lst.append(lnode_dic)
            lnode_dic = {}
            lnodeid_lst = []
        if '-->--> node_sn:' in line:
            node_id = line.split(',')[1].split(':')[-1].strip()
            lnode_dic['node_id'] = node_id
        if '-->-->--> lnodeid:' in line:
            lnodeid = line.split(',')[0].split(':')[-1].strip()
            lnodeid_lst.append(lnodeid)
            lnode_dic['lnode_id'] = lnodeid_lst

    for lnode_dic in node_lnode_lst:
        node_id = lnode_dic['node_id']
        lnode_lst = lnode_dic['lnode_id']
        for lnode_id in lnode_lst:
            cmd = '/home/parastor/tools/nWatch -i %s -t oPara -c oPara#vmgr_flattennr_dump -a "vmgrid=%s"' \
                  % (node_id, lnode_id)
            rc, stdout = run_command(cmd)
            if (0 != rc) or ('failed' in stdout) or ('support' in stdout):
                logging.warn("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
                return False
            vset_num = stdout.strip().split('\n')[-1].split()[2]
            try:
                if int(vset_num) != 0:
                    return False
                else:
                    continue
            except Exception, e:
                logging.error("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
                raise Exception("Error msg is %s" % e)
    logging.info("The current environment all vset is flatten")
    return True


def run_check_ds():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 检查是否所有ds是否提供服务
    :return: 
    """
    node_ids = get_nodes_id()
    for node_id in node_ids:
        cmd = '/home/parastor/tools/nWatch -i %s -t oStor -c oStor#get_basicinfo' % node_id
        rc, stdout = run_command(cmd)
        if 0 != rc or 'failed' in stdout.strip().splitlines()[0]:
            logging.warn("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
            return False
        else:
            stdout_lst = stdout.strip().split('\n')
            for line in stdout_lst:
                if 'provide serv' in line:
                    flag = line.split(':')[-1].strip()
                    try:
                        if 1 != int(flag):
                            return False
                    except Exception, e:
                        logging.error("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
                        raise Exception("Error msg is %s" % e)
    logging.info("The current environment all ds service is OK")
    return True


def run_check_jnl():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 检查jnl是否正常
    :return: 
    """
    mgr_node_id_lst = get_mgr_node_ids()
    mgr_node_id = random.choice(mgr_node_id_lst)
    cmd = '/home/parastor/tools/nWatch -t oRole -i %s -c oRole#rolemgr_master_dump' % mgr_node_id
    rc, stdout = run_command(cmd)
    if 0 != rc or 'failed' in stdout.strip().splitlines()[0]:
        logging.error("Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
        return False
    master_node_id = stdout.split(':')[-1].strip()
    cmd1 = '/home/parastor/tools/nWatch -t oRole -i %s -c oRole#rolemgr_slaveready_dump' % master_node_id
    rc, stdout = run_command(cmd1)
    if 0 != rc or 'failed' in stdout.strip().splitlines()[0]:
        logging.error("Execute command: \"%s\" failed. \nstdout: %s" % (cmd1, stdout))
        return False
    stdout_lst = stdout.strip().split('\n')
    for line in stdout_lst:
        if 'nodeid' in line and 'is_takeoverable' in line:
            node_id_tmp = line.split()[-2].split(':')[-1].rstrip(',')
            takeoverable = line.split()[-1].split(':')[-1].strip()
            if takeoverable != '1':
                logging.error("node %s jnl is not normal" % node_id_tmp)
                return False
    logging.info("The current environment all jnl service is OK")
    return True


def check_service_hook(func, info):
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 检查服务是否正常的钩子函数
    :param func:  具体检查服务状态的函数
    :param info:  提示的信息
    :return: 
    """
    start_time = time.time()
    while True:
        time.sleep(20)
        rc = func()
        if rc:
            break
        exist_time = int(time.time() - start_time)
        m, s = divmod(exist_time, 60)
        h, m = divmod(m, 60)
        if rc is False:
            time_str = "%s %dh:%dm:%ds" % (info, h, m, s)
            logging.info(time_str)
    return


def check_service(waitflag=True):
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 检查系统服务是否正常
    :return: 
    """
    """检查core"""
    check_core()

    """检查坏对象"""
    check_badobj(waitflag)

    """检查vset"""
    check_service_hook(run_check_vset, "vset exist")

    """检查ds"""
    check_service_hook(run_check_ds, "ds don't provide service")

    """检查jnl"""
    check_service_hook(run_check_jnl, "jnl is not OK")
    return


"""****************************** excle 相关操作 ******************************"""


def add_result_excel(row_idx, col_idx, result):
    """
    :author:        baoruobing
    :date  :        2018.08.15
    :description:   在excel中添加用例执行结果
    :param row_idx: 行索引
    :param col_idx: 列索引
    :param result:  用例执行结果
    :return: 
    """
    excel = openpyxl.load_workbook(Case_File)
    sheet = excel['case_list']
    sheet.cell(row=row_idx, column=col_idx, value=result)
    excel.save(Case_File)


def analysis_excel():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 解析用例文件,excel,生成要运行的用例列表
    :return: 
    """
    case_info_dic = {}
    excel = openpyxl.load_workbook(Case_File)
    sheet = excel['case_list']
    for row in sheet.iter_rows(min_row=2):
        fault_id_lst = []
        case_info_dic = {}
        for cell in row:
            cell_1_value = sheet.cell(row=1, column=cell.col_idx).value
            if u'case_num' in cell_1_value:
                case_info_dic['casename'] = cell.value
            elif u'次数' in cell_1_value:
                case_info_dic['nums'] = int(cell.value)
            elif u'故障' in cell_1_value:
                fault_id_lst.append(int(cell.value))
            elif u'节点数量' in cell_1_value:
                case_info_dic['nodenum'] = int(cell.value)
            elif u'节点冗余度' in cell_1_value:
                case_info_dic['node_parity_num'] = int(cell.value)
            elif u'磁盘冗余度' in cell_1_value:
                case_info_dic['disk_parity_num'] = int(cell.value)
            elif u'时间间隔' in cell_1_value:
                case_info_dic['interval'] = cell.value
            elif u'运行结果' in cell_1_value:
                case_info_dic['result_col_idx'] = cell.col_idx
        case_info_dic['row'] = cell.row
        case_info_dic['fault'] = fault_id_lst
        sheet.cell(row=case_info_dic['row'], column=case_info_dic['result_col_idx']).value = None
        if case_info_dic['nums'] != 0:
            Case_Lst.append(case_info_dic)
    excel.save(Case_File)

"""****************************** 参数解析 ******************************"""


def arg_analysis():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 参数解析
    :return: 
    """
    global NUMBERS
    global NODE_IP_LST
    global NODE_DATA_IP_LST
    global CORESTOP
    usage = "usage: %prog [options] arg1 arg2 arg3"
    version = "%prog 4.8"
    parser = OptionParser(usage=usage, version=version)
    parser.add_option("-i", "--ip",
                      type="str",
                      dest="mgrip",
                      default=None,
                      help="Required:True   Type:str   Help:MGR node IP")

    parser.add_option("-c", "--corestop",
                      action="store_true",
                      dest="core",
                      help="Required:False   Type:bool  Help:core stop. If you want to stop when there have cores, "
                           "you need to configure this parameter")

    parser.add_option("-n", "--numbers",
                      type="int",
                      dest="num",
                      default=10000,
                      help="Required:False   Type:int  Help:fault execution nums, default: %default, range is 1-10000")

    options, args = parser.parse_args()

    """检查-i参数"""
    if options.mgrip is None:
        parser.error("-i must be configed!!!")
    else:
        """检查ip的正确性"""
        if check_ip(options.mgrip) is False:
            parser.error("-i the ip format is incorrent!!!")
        mgr_ip = options.mgrip
        NODE_IP_LST = get_nodes_ips_by_ip(mgr_ip)
        NODE_DATA_IP_LST = get_nodes_data_ip_by_ip(mgr_ip)

        """如果本节点在集群内则报错"""
        def _check_localnode_in_ps(node_ip_lst):
            cmd = 'ip addr | grep "inet "'
            rc, stdout = command(cmd)
            if 0 != rc:
                raise Exception(
                    "Execute command: \"%s\" failed. \nstdout: %s" % (cmd, stdout))
            lines = stdout.strip().split('\n')
            for line in lines:
                ip = line.split()[1].split('/')[0]
                if ip in node_ip_lst:
                    return True
            return False

        if _check_localnode_in_ps(NODE_IP_LST):
            parser.error("If the local node in the parastor, please don't enter -i or --ip")

    """检查-c参数"""
    if options.core is True:
        CORESTOP = True
    else:
        CORESTOP = False

    """检查-n参数"""
    if options.num < 1 or options.num > 10000:
        parser.error("the range of -n or --num is 1-10000")
    NUMBERS = options.num
    return


"""****************************** 运行用例 ******************************"""


def run_func(func, *args):
    try:
        func(*args)
    except Exception:
        logging.error("", exc_info=1)
        traceback.print_exc()
        sys.exit(1)


def check_fault_conflict(parameter_lst):
    """
    :author:              baoruobing
    :date  :              2018.08.15
    :description:         参数解析
    :param parameter_lst: (list)[{'fault_id':1, 'node_ip':'10.2.40.1', 'disk_name':'sdb'},
                                 {'fault_id':2, 'node_ip':'10.2.40.1', 'disk_name':'sdc'}]
    :return: 
    """
    if parameter_lst[0]['fault_id'] in Fault_Type['node'] or parameter_lst[1]['fault_id'] in Fault_Type['node']:
        if parameter_lst[0]['node_ip'] == parameter_lst[1]['node_ip']:
            return False
    if parameter_lst[0]['fault_id'] in Fault_Type['disk'] and parameter_lst[1]['fault_id'] in Fault_Type['disk']:
        if parameter_lst[0]['node_ip'] == parameter_lst[1]['node_ip'] and \
                        parameter_lst[0]['disk_name'] == parameter_lst[1]['disk_name']:
            return False
    return True


def get_case_func(fault_id):
    """
    :author:         baoruobing
    :date  :         2018.08.15
    :description:    获取故障的相关函数和信息
    :param fault_id: 故障case id
    :return: 1:获取节点和磁盘函数; 2:执行故障函数; 3:故障提示信息
    """
    fault_func = {1:  [None, None, None],
                  2:  [None, None, None],
                  3:  [None, None, None],
                  4:  [None, None, None],
                  5:  [None, None, None],
                  6:  [None, None, None],
                  7:  [None, None, None],
                  8:  [None, None, None],
                  9:  [None, None, None],
                  10: [None, None, None],
                  11: [None, None, None],
                  12: [None, None, None],
                  13: [None, None, None],
                  14: [None, None, None],
                  15: [get_mgr_node,                   fault_half_net,         "管理节点网络故障1/2"],
                  16: [get_mgr_node,                   fault_all_net,          "管理节点网络故障全部"],
                  17: [get_no_mgr_node,                fault_half_net,         "非管理节点网络故障1/2"],
                  18: [get_no_mgr_node,                fault_all_net,          "非管理节点网络故障全部"],
                  19: [None, None, None],
                  20: [None, None, None],
                  21: [get_one_client_node,            fault_client_half_net,  "客户端网络故障1/2"],
                  22: [get_one_client_node,            fault_client_all_net,   "客户端网络故障全部"],
                  23: [get_lmos_node_and_meta_disk,    fault_disk_no_rebuild,  "主mos共享盘故障拔出"],
                  24: [get_no_lmos_node_and_meta_disk, fault_disk_no_rebuild,  "从mos共享盘故障拔出"],
                  25: [get_random_node_data_disk,      fault_disk_no_rebuild,  "数据盘故障拔出不重建"],
                  26: [get_random_node_data_disk,      fault_disk_rebuild,     "数据盘故障拔出重建"],
                  27: [None, None, None],
                  28: [None, None, None],
                  29: [get_master_zk_node,             fault_master_zk_pro,    "zk进程换主"],
                  30: [get_master_orole_node,          fault_orole_pro,        "oRole主进程故障"],
                  31: [get_no_master_orole_node,       fault_orole_pro,        "oRole从进程故障"],
                  32: [get_random_node,                fault_ostor_pro,        "oStor进程故障"],
                  33: [get_lmos_node,                  fault_opara_pro,        "oPara主进程故障"],
                  34: [get_no_lmos_node,               fault_opara_pro,        "oPara从进程故障"],
                  35: [get_master_ojob_node,           fault_ojob_pro,         "oJob主进程故障"],
                  36: [get_no_master_ojob_node,        fault_ojob_pro,         "oJob从进程故障"],
                  37: [get_master_mgr_node,            fault_ojmgs_pro,        "oJmgs主进程故障"],
                  38: [get_no_master_mgr_node,         fault_ojmgs_pro,        "oJmgs从进程故障"],
                  39: [get_random_node,                fault_omgcd_pro,        "集群节点oMgcd故障"],
                  40: [get_one_client_node,            fault_omgcd_client_pro, "客户端节点oMgcd故障"]}
    return fault_func[fault_id][0], fault_func[fault_id][1], fault_func[fault_id][2]


def run_fault_case(vdb_fail_flag):
    """
    :author:              baoruobing
    :date  :              2018.08.15
    :description:         获取故障的相关函数和信息
    :param vdb_fail_flag: 运行故障用例
    :return: 
    """
    """如果vdbench失败了，则退出"""
    if vdb_fail_flag.value == 1:
        raise Exception("vdbench failed!!!")

    logging.info("*********** the fault operation beginning ***********")

    """检查服务状态"""
    check_service(waitflag=False)

    num = 0
    for i in range(NUMBERS):
        for fault_case in Case_Lst:
            """如果vdbench失败了，则退出"""
            if vdb_fail_flag.value == 1:
                raise Exception("vdbench failed!!!")
            get_fault_info_func1, make_fault_func1, fault_info1 = get_case_func(fault_case['fault'][0])
            get_fault_info_func2, make_fault_func2, fault_info2 = get_case_func(fault_case['fault'][1])
            logging.info("***************************** %s begin *****************************"
                         % fault_case['casename'])
            logging.info("将要执行下面两个故障")
            logging.info(fault_info1)
            logging.info(fault_info2)

            def _set_result(result):
                fault_case['result'] = result
                add_result_excel(fault_case['row'], fault_case['result_col_idx'], fault_case['result'])

            """检查环境是否满足用例"""
            if not check_case_layout(fault_case['nodenum'], fault_case['node_parity_num'],
                                     fault_case['disk_parity_num']):
                _set_result('skip')
                logging.warn("Environment does not meet the minimum conditions of use cases!")
                logging.warn("case skip")
                continue

            """获取node_id, disk_name"""
            for i in range(5):
                parameter_lst = []
                for fault_id in fault_case['fault']:
                    parameter_dic = {}
                    parameter_dic['fault_id'] = fault_id
                    node_ip, disk_name = get_case_func(fault_id)[0]()
                    parameter_dic['node_ip'] = node_ip
                    parameter_dic['disk_name'] = disk_name
                    parameter_lst.append(parameter_dic)

                """检查是否可以做故障"""
                if check_fault_conflict(parameter_lst):
                    break
            else:
                _set_result('skip')
                logging.warn("Unable to select faulty node or disk 5 times!")
                logging.warn("case skip")
                continue

            def _get_interval(interval):
                if '-' not in interval:
                    wait_time = int(interval)
                else:
                    min_time = int(interval.split('-')[0])
                    max_time = int(interval.split('-')[-1])
                    wait_time = random.randint(min_time, max_time)
                return wait_time

            for i in range(fault_case['nums']):
                """如果vdbench失败了，则退出"""
                if vdb_fail_flag.value == 1:
                    logging.info("***************************** %s failed *****************************\n"
                                 % fault_case['casename'])
                    _set_result('failed')
                    return
                """并行制造故障"""
                num += 1
                fault_id1 = parameter_lst[0]['fault_id']
                fault_id2 = parameter_lst[1]['fault_id']
                logging.info('***************************** the %d fault begin ******************************' % num)
                try:
                    t1 = Process(target=run_func,
                                 args=(get_case_func(fault_id1)[1],
                                       parameter_lst[0]['node_ip'],
                                       parameter_lst[0]['disk_name']))
                    t2 = Process(target=run_func,
                                 args=(get_case_func(fault_id2)[1],
                                       parameter_lst[1]['node_ip'],
                                       parameter_lst[1]['disk_name']))

                    t1.daemon = True
                    t2.daemon = True

                    t1.start()
                    wait_time = _get_interval(fault_case['interval'])
                    logging.info('wait %ss' % wait_time)
                    time.sleep(wait_time)
                    t2.start()

                    t1.join()
                    t2.join()

                    if t1.exitcode != 0 or t2.exitcode != 0:
                        _set_result('failed')
                        sys.exit(1)

                    """检查服务状态"""
                    check_service()
                except Exception:
                    logging.info("***************************** %s failed *****************************\n"
                                 % fault_case['casename'])
                    _set_result('failed')
                    logging.error("", exc_info=1)
                    traceback.print_exc()
                    sys.exit(1)

                """如果vdbench失败了，则退出"""
                if vdb_fail_flag.value == 1:
                    logging.info("***************************** %s failed *****************************\n"
                                 % fault_case['casename'])
                    _set_result('failed')
                    sys.exit(1)

                logging.info("case success!!!")
                logging.info('***************************** the %d fault finish ******************************' % num)
            _set_result('success')
            logging.info("***************************** %s success *****************************\n"
                         % fault_case['casename'])


def run_case():
    """两个进程分别跑vdbench业务和故障"""
    mgr = Manager()
    vdb_fail_flag = mgr.Value('i', 0)

    pro_vdb = Process(target=run_vdbench, args=(vdb_fail_flag,))
    pro_fault = Process(target=run_fault_case, args=(vdb_fail_flag,))

    pro_vdb.daemon = True

    pro_vdb.start()
    time.sleep(20)
    pro_fault.start()

    time.sleep(10)
    while True:
        time.sleep(10)
        if not (pro_vdb.is_alive() or pro_fault.is_alive()):
            """两个进程都退出了，则退出"""
            if pro_fault.exitcode == 0:
                logging.info("test success!!!")
            break
        elif not pro_fault.is_alive():
            """运行故障进程退出了，vdbench直接强退"""
            os.killpg(os.getpgid(pro_vdb.pid), signal.SIGTERM)
            if pro_fault.exitcode == 0:
                logging.info("test success!!!")
            break

    if pro_fault.exitcode != 0:
        logging.error("test failed!!!")


def main():
    """
    :author:      baoruobing
    :date  :      2018.08.15
    :description: 主函数
    :return: 
    """
    """参数解析"""
    arg_analysis()
    """初始化日志"""
    log_init()
    """解析excel,获取用例列表"""
    analysis_excel()
    """用例执行"""
    run_case()

if __name__ == '__main__':
    run_func(main)